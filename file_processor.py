import csv
import io
import re
import warnings
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from tiktok_stats import fetch_many_video_data


URL_HEADER_NAMES = {
    "url",
    "link",
    "links",
    "tiktok",
    "tiktokurl",
    "tiktoklink",
    "tiktok_url",
    "tiktok_link",
    "video",
    "videourl",
    "video_url",
}
STAT_HEADERS = [
    "Caption",
    "View",
    "Like",
    "Comment",
    "Share",
    "Save",
]
COUNT_HEADERS = {"View", "Like", "Comment", "Share", "Save"}
TIKTOK_URL_RE = re.compile(r"https?://(?:www\.|m\.)?tiktok\.com/|https?://v[mt]\.tiktok\.com/")


@dataclass
class TableData:
    headers: list[str]
    rows: list[list[Any]]
    url_col: int
    source_type: str


def is_tiktok_url(value: Any) -> bool:
    return bool(TIKTOK_URL_RE.search(str(value or "").strip()))


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _find_url_col(headers: list[Any], rows: list[list[Any]]) -> int:
    for index, header in enumerate(headers):
        if _normalize_header(header) in URL_HEADER_NAMES:
            return index

    for row in rows[:10]:
        for index, value in enumerate(row):
            if is_tiktok_url(value):
                return index

    return 0


def _looks_like_header(first_row: list[Any], remaining_rows: list[list[Any]]) -> bool:
    if any(_normalize_header(value) in URL_HEADER_NAMES for value in first_row):
        return True
    if any(is_tiktok_url(value) for value in first_row):
        return False
    return bool(remaining_rows)


def _pad_row(row: list[Any], size: int) -> list[Any]:
    if len(row) >= size:
        return row
    return row + [""] * (size - len(row))


def _format_count(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{int(value):,}"
    except (TypeError, ValueError):
        return str(value)


def parse_upload(filename: str, content: bytes) -> TableData:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return parse_csv(content)
    if suffix in {".xlsx", ".xlsm"}:
        if not zipfile.is_zipfile(io.BytesIO(content)):
            try:
                return parse_csv(content)
            except UnicodeDecodeError as exc:
                raise ValueError(
                    "File có đuôi .xlsx nhưng không phải định dạng Excel hợp lệ. "
                    "Vui lòng export lại dưới dạng .xlsx hoặc .csv."
                ) from exc
        return parse_xlsx(content)
    raise ValueError("Chỉ hỗ trợ file .csv, .xlsx hoặc .xlsm.")


def parse_csv(content: bytes) -> TableData:
    if content.startswith((b"\xff\xfe", b"\xfe\xff")):
        text = content.decode("utf-16")
    else:
        text = content.decode("utf-8-sig")
    rows = [row for row in csv.reader(io.StringIO(text)) if any(cell.strip() for cell in row)]
    if rows and rows[0] and rows[0][0].strip().lower().startswith("sep="):
        rows = rows[1:]
    if not rows:
        raise ValueError("File CSV đang trống.")

    first_row = rows[0]
    remaining = rows[1:]
    if _looks_like_header(first_row, remaining):
        headers = [str(value or "").strip() or f"column_{index + 1}" for index, value in enumerate(first_row)]
        data_rows = remaining
    else:
        headers = [f"column_{index + 1}" for index in range(len(first_row))]
        data_rows = rows

    width = max(len(headers), *(len(row) for row in data_rows), 1)
    headers = _pad_row(headers, width)
    data_rows = [_pad_row(row, width) for row in data_rows]
    url_col = _find_url_col(headers, data_rows)

    return TableData(headers=headers, rows=data_rows, url_col=url_col, source_type="csv")


def parse_xlsx(content: bytes) -> TableData:
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception:
        return parse_xlsx_xml(content)
    sheet = workbook.active
    if sheet is None:
        return parse_xlsx_xml(content)
    rows = []
    for row in sheet.iter_rows():
        values = []
        for cell in row:
            value = cell.value
            hyperlink_target = cell.hyperlink.target if cell.hyperlink else None
            if hyperlink_target and is_tiktok_url(hyperlink_target):
                value = hyperlink_target
            values.append(value)
        if any(value not in (None, "") for value in values):
            rows.append(values)
    if not rows:
        raise ValueError("File Excel đang trống.")

    first_row = rows[0]
    remaining = rows[1:]
    if _looks_like_header(first_row, remaining):
        headers = [str(value or "").strip() or f"column_{index + 1}" for index, value in enumerate(first_row)]
        data_rows = remaining
    else:
        headers = [f"column_{index + 1}" for index in range(len(first_row))]
        data_rows = rows

    width = max(len(headers), *(len(row) for row in data_rows), 1)
    headers = _pad_row(headers, width)
    data_rows = [_pad_row(row, width) for row in data_rows]
    url_col = _find_url_col(headers, data_rows)

    return TableData(headers=headers, rows=data_rows, url_col=url_col, source_type="xlsx")


def _cell_column_index(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
    index = 0
    for letter in letters:
        index = index * 26 + (ord(letter) - ord("A") + 1)
    return max(index - 1, 0)


def _xml_text(element: ElementTree.Element) -> str:
    return "".join(element.itertext())


def parse_xlsx_xml(content: bytes) -> TableData:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.iter():
                if item.tag.endswith("}si") or item.tag == "si":
                    shared_strings.append(_xml_text(item))

        sheet_names = [name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")]
        if not sheet_names:
            raise ValueError("File Excel không có worksheet hợp lệ.")

        sheet_root = ElementTree.fromstring(archive.read(sheet_names[0]))
        rows: list[list[Any]] = []
        for row_el in sheet_root.iter():
            if not (row_el.tag.endswith("}row") or row_el.tag == "row"):
                continue
            row_values: list[Any] = []
            for cell_el in row_el:
                if not (cell_el.tag.endswith("}c") or cell_el.tag == "c"):
                    continue
                col_index = _cell_column_index(cell_el.attrib.get("r", "A"))
                while len(row_values) <= col_index:
                    row_values.append("")
                cell_type = cell_el.attrib.get("t")
                value_text = ""
                for child in cell_el:
                    if child.tag.endswith("}v") or child.tag == "v":
                        value_text = child.text or ""
                        break
                    if child.tag.endswith("}is") or child.tag == "is":
                        value_text = _xml_text(child)
                        break
                if cell_type == "s" and value_text != "":
                    try:
                        value = shared_strings[int(value_text)]
                    except (ValueError, IndexError):
                        value = value_text
                else:
                    value = value_text
                row_values[col_index] = value
            if any(value not in (None, "") for value in row_values):
                rows.append(row_values)

    if not rows:
        raise ValueError("File Excel đang trống.")

    first_row = rows[0]
    remaining = rows[1:]
    if _looks_like_header(first_row, remaining):
        headers = [str(value or "").strip() or f"column_{index + 1}" for index, value in enumerate(first_row)]
        data_rows = remaining
    else:
        headers = [f"column_{index + 1}" for index in range(len(first_row))]
        data_rows = rows

    width = max(len(headers), *(len(row) for row in data_rows), 1)
    headers = _pad_row(headers, width)
    data_rows = [_pad_row(row, width) for row in data_rows]
    url_col = _find_url_col(headers, data_rows)
    return TableData(headers=headers, rows=data_rows, url_col=url_col, source_type="xlsx")


async def enrich_table(
    table: TableData,
    *,
    browser: str,
    headless: bool,
    max_rows: int,
) -> TableData:
    if len(table.rows) > max_rows:
        raise ValueError(f"Batch limit is {max_rows} rows. Split the file and try again.")

    urls: list[str] = []
    lookup: dict[int, int] = {}
    for row_index, row in enumerate(table.rows):
        url = str(row[table.url_col] if table.url_col < len(row) else "").strip()
        if is_tiktok_url(url):
            lookup[row_index] = len(urls)
            urls.append(url)

    results = await fetch_many_video_data(urls, browser=browser, headless=headless) if urls else []
    headers = table.headers + STAT_HEADERS
    enriched_rows: list[list[Any]] = []

    for row_index, row in enumerate(table.rows):
        output_row = list(row)
        result_index = lookup.get(row_index)
        if result_index is None:
            output_row.extend(["", "", "", "", "", ""])
        else:
            result = results[result_index]
            if result["error"]:
                output_row.extend(["", "", "", "", "", ""])
            else:
                data = result["data"]
                stats = data.get("stats", {})
                output_row.extend(
                    [
                        data.get("description"),
                        _format_count(stats.get("view")),
                        _format_count(stats.get("like")),
                        _format_count(stats.get("comment")),
                        _format_count(stats.get("share")),
                        _format_count(stats.get("collect")),
                    ]
                )
        enriched_rows.append(output_row)

    return TableData(
        headers=headers,
        rows=enriched_rows,
        url_col=table.url_col,
        source_type=table.source_type,
    )


def results_to_link_table(urls: list[str], results: list[dict[str, Any]]) -> TableData:
    rows: list[list[Any]] = []
    for url, result in zip(urls, results, strict=False):
        if result["error"]:
            rows.append([url, "", "", "", "", "", ""])
            continue

        data = result["data"]
        stats = data.get("stats", {})
        rows.append(
            [
                url,
                data.get("description"),
                _format_count(stats.get("view")),
                _format_count(stats.get("like")),
                _format_count(stats.get("comment")),
                _format_count(stats.get("share")),
                _format_count(stats.get("collect")),
            ]
        )

    return TableData(
        headers=["Url"] + STAT_HEADERS,
        rows=rows,
        url_col=0,
        source_type="csv",
    )


def urls_from_table(table: TableData) -> list[str]:
    urls: list[str] = []
    for row in table.rows:
        url = str(row[table.url_col] if table.url_col < len(row) else "").strip()
        if is_tiktok_url(url):
            urls.append(url)
    return urls


def table_to_csv(table: TableData) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    output.write("sep=,\r\n")
    writer.writerow(table.headers)
    for row in table.rows:
        writer.writerow(
            [
                f'="{value}"' if header in COUNT_HEADERS and value not in (None, "") else value
                for header, value in zip(table.headers, row, strict=False)
            ]
        )
    return output.getvalue().encode("utf-16")


def table_to_xlsx(table: TableData) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TikTok Stats"
    sheet.append(table.headers)
    for row in table.rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        column_index = column_cells[0].column
        values = [str(cell.value or "") for cell in column_cells[:100]]
        width = min(max(max((len(value) for value in values), default=10) + 2, 10), 60)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
